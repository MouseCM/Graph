import matplotlib.pyplot as plt
import numpy as np
import openpyxl as open

def process(file_pos, col_str, start_str):
    file = file_pos
    col = int(col_str)
    start = int(start_str)

    wb = open.load_workbook(file)
    ws = wb.active

    max_row = ws.max_row

    x = []
    y = []

    for i in range(start, max_row):
        cell = ws.cell(i, col)
        if(cell.value not in y):
            if cell.value == None:
                continue
            y.append(cell.value)
            x.append(0)

            pos = len(y) - 1
            x[pos] += 1
        else:
            pos = y.index(cell.value)
            x[pos] += 1
            
    try:
        quickSort(x, y, 0, len(y)-1)        
    except:
        pass

    plt.bar(y, x)
    plt.show()

def partition(array1, array2, low, high):
    pivot = array2[high]
    i = low - 1

    for j in range(low, high):
        if array2[j] <= pivot:
            i = i + 1
            (array2[i], array2[j]) = (array2[j], array2[i])
            (array1[i], array1[j]) = (array1[j], array1[i])

    (array1[i + 1], array1[high]) = (array1[high], array1[i + 1])
    (array2[i + 1], array2[high]) = (array2[high], array2[i + 1])

    return i + 1

def quickSort(array1, array2, low, high):
    if low < high:
        pi = partition(array1, array2, low, high)
        quickSort(array1, array2, low, pi - 1)
        quickSort(array1, array2, pi + 1, high)

